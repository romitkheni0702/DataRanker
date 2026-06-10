# services/ranker.py — Stage 3: Score and rank companies by KPI template

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill

from core.config import KPI_HEADER_ROW, KPI_SHEET_NAME, METRIC_SCORE_COLOR, RANK_COLOR


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _score_kpi(
    subset: pd.DataFrame,
    kpi: str,
    weight: float,
    direction: str,
) -> tuple[pd.DataFrame, str]:
    """
    Compute rank, percentile slab, and weighted metric score for one KPI.

    Direction semantics (from the KPI Library "Higher/Lower Better" column):
      - "Higher" → a larger value is better → rank 1 goes to the largest value
                   → pandas rank with ascending=False
      - "Lower"  → a smaller value is better → rank 1 goes to the smallest value
                   → pandas rank with ascending=True

    The _Rank column uses the same convention so rank 1 always means "best".
    The percentile slab is derived the same way so that the best company always
    lands in the top slab (100) and the worst in the bottom slab (10), regardless
    of direction.

    Returns the mutated subset and the name of the new metric-score column.
    """
    higher_is_better: bool = direction.strip().lower() == "higher"

    # Rank 1 = best company for this KPI
    #   "Higher" → largest value is best → ascending=False so rank 1 → largest
    #   "Lower"  → smallest value is best → ascending=True  so rank 1 → smallest
    subset[f"{kpi}_Rank"] = subset[kpi].rank(ascending=not higher_is_better)

    # Percentile slab: best company → 100th percentile, worst → 10th percentile
    #   pandas rank(pct=True, ascending=True)  → largest value gets pct≈1.0
    #   pandas rank(pct=True, ascending=False) → smallest value gets pct≈1.0
    #   "Higher" → want largest  to score highest → ascending=True
    #   "Lower"  → want smallest to score highest → ascending=False
    raw_pct = subset[kpi].rank(pct=True, ascending=higher_is_better) * 100
    subset[f"{kpi}_Percentile_Slab"] = np.ceil(raw_pct / 10) * 10

    score_col = f"{kpi}_Metric_Score"
    subset[score_col] = subset[f"{kpi}_Percentile_Slab"] * weight

    return subset, score_col


def _apply_cell_fills(sheet, columns: list[str], n_rows: int) -> None:
    """Colour metric-score and company-rank columns in the given sheet."""
    metric_fill = _make_fill(METRIC_SCORE_COLOR)
    rank_fill = _make_fill(RANK_COLOR)

    for col_idx, col_name in enumerate(columns, 1):
        if "_Metric_Score" in col_name:
            fill = metric_fill
        elif col_name == "Company_Rank":
            fill = rank_fill
        else:
            continue

        for row_idx in range(2, n_rows + 2):
            sheet.cell(row=row_idx, column=col_idx).fill = fill


# ── Public entry point ────────────────────────────────────────────────────────

def run_ranking(mapped_path: str, kpi_lib_path: str, out_path: str) -> None:
    """
    For every KPI template found in the mapped CSV, score and rank companies
    using the KPI library weights, then write each template to its own sheet.

    Args:
        mapped_path:  Path to the Stage-2 mapped CSV.
        kpi_lib_path: Path to the KPI library .xlsx workbook.
        out_path:     Path where the final ranked .xlsx will be written.
    """
    df = pd.read_csv(mapped_path)
    kpi_lib = pd.read_excel(kpi_lib_path, sheet_name=KPI_SHEET_NAME, header=KPI_HEADER_ROW)
    kpi_lib.columns = kpi_lib.columns.str.strip()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for template in df["KPI_Template"].dropna().unique():
            _write_template_sheet(writer, df, kpi_lib, template)


def _write_template_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    kpi_lib: pd.DataFrame,
    template: str,
) -> None:
    """Score, rank, and write one template's companies to its own Excel sheet."""
    subset = df[df["KPI_Template"] == template].copy()
    metrics = kpi_lib[kpi_lib["Template"] == template]

    score_cols: list[str] = []
    for _, row in metrics.iterrows():
        kpi = row["KPI"]
        weight = row["Weight %"] / 100
        direction = row["Higher/Lower Better"]

        # FIX 3: A column may exist but be entirely NaN because formatter.py
        # pre-created it as a placeholder (EXTRA_COLUMNS). Scoring an all-NaN
        # series produces all-NaN _Rank / _Percentile_Slab / _Metric_Score cols.
        # Skip any KPI whose values are all missing.
        if kpi in subset.columns and subset[kpi].notna().any():
            subset, score_col = _score_kpi(subset, kpi, weight, direction)
            score_cols.append(score_col)

    subset["Total_Final_Score"] = subset[score_cols].sum(axis=1)
    subset["Company_Rank"] = subset["Total_Final_Score"].rank(
        ascending=False, method="min"
    )

    sheet_name = str(template)[:31]
    subset.to_excel(writer, sheet_name=sheet_name, index=False)

    _apply_cell_fills(writer.sheets[sheet_name], list(subset.columns), len(subset))
